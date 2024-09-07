from openpyxl import *
import numpy as np
import pandas as pd
import smtplib
import imaplib
from email import encoders
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
try:
    wb = load_workbook("attendance_tracker.xlsx")
    ws = wb.active
except :
    print("File not found")
    wb = Workbook()
    ws = wb.active
    ws.title = "Main sheet"
while True :
    print("Press 0 to terminate the program\nPress 1 to insert a value\nPress 2 to remove a value\nPress 3 to show all the values\nPress 4 if you want to record a student's leave")
    a = int(input(""))
    if (a==0) :
        break
    elif (a==1):
        try :
            name=str(input("Enter the Reg.No of student : "))
            if not name :
                raise NameError("Reg.No is compulsory")
        except NameError as e:
            print(e)
            continue
        try :
            mail=str(input("Enter the address of mail : "))
            if not mail :
                raise NameError("Mail is compulsory")
        except NameError as e:
            print(e)
            continue
        try :
            print("Please enter only integer values for the next column")
            number_of_leaves=int(input("Enter the number of leaves : "))
            if not number_of_leaves :
                raise NameError("number of leaves is compulsory")
            if number_of_leaves <= 0:
                raise ValueError("Number of leaves must be positive")
        except NameError as e:
            print(e)
            continue
        except ValueError as v:
            print(v)
            continue
        except TypeError as p :
            print(p)
            continue
        try :
            subject_code=str(input("Enter the subject code : "))
            if not subject_code :
                raise NameError("subject code is compulsory")
        except NameError as e:
            print(e)
            continue
        ws.append([name,mail,number_of_leaves,subject_code])
        wb.save("attendance_tracker.xlsx")
    elif (a == 2):
        try :
            print("Please enter the registration number to be removed: ", end="")
            x = str(input(""))
            n = []
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                if not any(row):
                    break
                n.append(row[0])
            if x in n:
                q = n.index(x)
                q = q + 1
                ws.delete_rows(q)
                wb.save("attendance_tracker.xlsx")
                print("Record removed successfully.")
            else:
                print(f"Reg.No {x} not found.")
        except FileNotFoundError as t :
            print(t)
    elif (a==3):
        try :
            n = []
            for row in ws.iter_rows(min_col=1, max_col=4, values_only=True):
                if not any(row):
                    break
                n.append(row)
            df=pd.DataFrame(n,columns=["Reg.No", "Mail", "number_of_leaves", "subject_code"])
            df.drop(index=0)
            print(df)
        except KeyError :
            print("We encountered a few problems")
        except FileNotFoundError :
            print("File not found")
    elif (a == 4):
        reg_no = str(input("Enter the reg.no of the student : "))
        subject_code = str(input("Enter the subject code : "))
        try:
            n = []
            for row in ws.iter_rows(min_col=1, max_col=4, values_only=True):
                if not any(row):
                    break
                n.append(row)
            df = pd.DataFrame(n, columns=["Reg.No", "Mail", "number_of_leaves", "subject_code"])
            a = list(df["Reg.No"])
            b = list(df["subject_code"])
            if reg_no not in a:
                print(f"{reg_no} not found")
                continue
            if subject_code not in b:
                print(f"{subject_code} not found for {reg_no}")
                continue
            for i in range(len(a)):
                if a[i] == reg_no and b[i] == subject_code:
                    i = i + 1
                    cell_value = ws.cell(row=i, column=3).value
                    if cell_value > 0:
                        ws.cell(row=i, column=3, value=cell_value - 1)
                        print(f"One leave deducted for {reg_no} in {subject_code}")
                        wb.save("attendance_tracker.xlsx")
                    else:
                        print(f"No leaves left for {reg_no} in {subject_code}")
                        server=smtplib.SMTP('smtp.gmail.com', 587)
                        server.starttls()
                        server.ehlo()
                        email_add="trackerattendance28@gmail.com"
                        password=str(input("Enter the password : "))
                        server.login(email_add,password)
                        msg=MIMEMultipart()
                        c='B'+str(i)
                        msg["From"]=f"{email_add}"
                        msg["To"]=f"{ws[c]}"
                        msg["Subject"]="Lack of Attendance"
                        message=f"This mail is sent to inform yourself and concerned parties that you have a lack of attendance in the subject with subject code {subject_code}."
                        msg.attach(MIMEText(message,'plain'))
                        text=msg.as_string()
                        server.sendmail(email_add,ws.cell(row=i,column=2).value,text)
                        server.quit()
        except KeyError:
            print("We encountered a few problems")
        except FileNotFoundError:
            print("File not found")
print("Program terminated")
